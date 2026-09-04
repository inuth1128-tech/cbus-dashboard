#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cambo Unisoll QC Dashboard - Data Extraction from Excel
각 시트별 정확한 구조로 처리
"""

import os
import json
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl이 설치되지 않았습니다.")
    print("다음 명령어로 설치하세요:")
    print("pip install openpyxl --break-system-packages")
    exit(1)

# ============ 경로 설정 ============
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(CURRENT_DIR)
DATA_FOLDER = os.path.join(PARENT_DIR, "Sewing Process(CBUS)")

if not os.path.exists(DATA_FOLDER):
    DATA_FOLDER = r"Y:\00 ALL REPORT\1) INSPECTION SUMMARY (QC & 3rd Party)\12)  Weekly_  Jun9~_ Top 3 Defect_ Improvement Project\Mr. Yang format\Sewing Process(CBUS)"

OUTPUT_FILENAME = "dashboard_data_enhanced.json"

# ============ Excel 컬럼 매핑 (각 시트별 정확한 구조) ============
SHEET_CONFIG = {
    'ENLINE': {
        'qty_col': 33, 'defects_col': 31, 'header_row': 3,
        'data_start': 4, 'data_end': 28, 'defect_cols': (6, 30),
        'line_col': 1, 'item_col': 5,  # A=Line, E=Item ID
    },
    'FINAL': {
        'qty_col': 27, 'defects_col': 25, 'header_row': 3,
        'data_start': 4, 'data_end': 27, 'defect_cols': (6, 24),
        'line_col': 1, 'item_col': 5,  # A=Line, E=Item ID
    },
    'FQC': {
        'qty_col': 5, 'defects_col': 6, 'header_row': 5,
        'data_start': 6, 'data_end': None, 'defect_cols': (8, 44),
        'line_col': 1, 'item_col': 5,  # A=Line, E=Item ID
    },
    'PQC 10%': {
        'qty_col': 7, 'defects_col': 8, 'header_row': 2,
        'data_start': 5, 'data_end': None, 'defect_cols': (13, 50),
        'line_col': 3, 'item_col': 4,  # C=Line, D=Item ID
    },
    'PQC 2.5%': {
        'qty_col': 7, 'defects_col': 8, 'header_row': 2,
        'data_start': 5, 'data_end': None, 'defect_cols': (13, 50),
        'line_col': 3, 'item_col': 4,  # C=Line, D=Item ID
    },
}

# 색상 팔레트
DEFAULT_COLORS = [
    "#1F4788","#FF6B6B","#4ECDC4","#45B7D1","#FFA07A",
    "#98D8C8","#F7DC6F","#BB8FCE","#85C1E2","#F8B8B8",
    "#B4E7FF","#FFD700","#FF69B4","#90EE90","#DEB887",
    "#87CEEB","#FFB347","#DDA0DD","#F0E68C","#20B2AA",
    "#FF8C00","#6495ED","#DC143C","#3CB371","#8B4513",
    "#FF4500","#228B22","#4169E1","#32CD32","#9932CC",
    "#CD5C5C","#2F4F4F","#00CED1","#9370DB","#3D95CE",
    "#6A5ACD","#00FA9A","#D2691E","#00688B","#8FBC8F",
    "#5F9EA0","#B0C4DE","#FFE4E1","#F5DEB3","#DBC295"
]

def clean_defect_type(text):
    """불량유형 이름 정제"""
    if not text:
        return None
    text = str(text).strip()
    text = text.replace('"', '').replace('\n', ' ').replace('\r', '')
    text = ' '.join(text.split())
    if not text or text == '':
        return None
    return text

def get_week_code(filename):
    """파일명에서 주차 코드 추출"""
    # 방법1: 괄호 안의 숫자 찾기
    match = re.search(r'\(([^)]*)\)', filename)
    if match:
        content = match.group(1)
        numbers = re.findall(r'\d{4}', content)
        if len(numbers) >= 2:
            return f"{numbers[0]}{numbers[1]}"
    
    # 방법2: 전체 파일명에서 4자리 숫자 찾기
    numbers = re.findall(r'\d{4}', filename)
    if len(numbers) >= 2:
        return f"{numbers[-2]}{numbers[-1]}"
    
    return None

def extract_item_id(item_str):
    """E열/D열에서 Item ID 추출: N031, F064, H069 등 (알파벳1개 + 숫자3개)
    S + 5자리(제조사 코드)는 제외"""
    if not item_str:
        return "Unknown"
    
    item_str = str(item_str).strip()
    
    # S + 5자리 제조사 코드 제거 (S02346 같은 패턴)
    cleaned = re.sub(r'S\d{5}', '', item_str)
    
    # 정제된 문자열에서 알파벳 + 3자리 숫자 추출
    match = re.search(r'[A-Z]\d{3}', cleaned)
    if match:
        return match.group()
    
    return "Unknown"

def find_data_end_row(sheet, config):
    """동적으로 데이터 끝 행 찾기"""
    if config['data_end'] is not None:
        return config['data_end']
    
    qty_col = config['qty_col']
    for row in range(config['data_start'], 500):
        cell_value = sheet.cell(row, qty_col).value
        if cell_value is None or cell_value == '':
            return row - 1
    return 100

def extract_sheet_data(sheet, sheet_name, config):
    """시트에서 데이터 추출"""
    print(f"  ✓ {sheet_name} 시트 처리 중...", end='', flush=True)
    
    qty_col = config['qty_col']
    defects_col = config['defects_col']
    header_row = config['header_row']
    data_start = config['data_start']
    data_end = find_data_end_row(sheet, config)
    defect_cols_range = config['defect_cols']
    line_col = config['line_col']
    item_col = config['item_col']
    
    # 불량 유형 헤더 추출
    defect_type_headers = {}
    for col in range(defect_cols_range[0], defect_cols_range[1] + 1):
        header_text = sheet.cell(header_row, col).value
        clean_name = clean_defect_type(header_text)
        if clean_name:
            defect_type_headers[col] = clean_name
    
    # 데이터 추출
    kpi_qty = 0
    kpi_defects = 0
    defect_counts = {}
    items_data = []
    
    for row in range(data_start, data_end + 1):
        try:
            qty = sheet.cell(row, qty_col).value or 0
            defects = sheet.cell(row, defects_col).value or 0
            line_num = sheet.cell(row, line_col).value
            item_str = sheet.cell(row, item_col).value
            
            try:
                qty = int(qty) if qty else 0
                defects = int(defects) if defects else 0
            except:
                continue
            
            if qty == 0 and defects == 0:
                continue
            
            kpi_qty += qty
            kpi_defects += defects
            
            # 불량 유형별 수량
            defect_detail = {}
            for col, defect_type in defect_type_headers.items():
                count = sheet.cell(row, col).value or 0
                try:
                    count = int(count) if count else 0
                except:
                    count = 0
                
                if count > 0:
                    defect_detail[defect_type] = count
                    defect_counts[defect_type] = defect_counts.get(defect_type, 0) + count
            
            # Item ID 추출
            item_id = extract_item_id(item_str)
            
            # Line 정리
            line_str = str(int(line_num)).strip() if isinstance(line_num, (int, float)) else str(line_num).strip() if line_num else "Unknown"
            
            items_data.append({
                "item": item_id,
                "line": line_str,
                "washing": "G/W",
                "qty": qty,
                "defects": defects,
                "defect_detail": defect_detail
            })
        except Exception as e:
            continue
    
    kpi_rate = round((kpi_defects / kpi_qty * 100), 2) if kpi_qty > 0 else 0
    
    print(f" ✓ ({len(items_data)}개 항목)")
    
    return {
        "kpi": {"qty": kpi_qty, "defects": kpi_defects, "rate": kpi_rate},
        "defects": defect_counts,
        "items": items_data
    }

def main():
    print("🔄 Cambo Unisoll QC Dashboard - Data Extraction")
    print("=" * 70)
    print(f"📂 데이터 폴더: {DATA_FOLDER}")
    
    # 데이터 폴더 확인
    if not os.path.exists(DATA_FOLDER):
        print(f"❌ 데이터 폴더를 찾을 수 없습니다!")
        print(f"   경로: {DATA_FOLDER}")
        return False
    
    # Excel 파일 찾기
    try:
        all_files = os.listdir(DATA_FOLDER)
        excel_files = [f for f in all_files if f.endswith('.xlsx')]
        excel_files = sorted([os.path.join(DATA_FOLDER, f) for f in excel_files])
    except Exception as e:
        print(f"❌ 폴더 읽기 실패: {e}")
        return False
    
    if not excel_files:
        print(f"❌ Excel 파일을 찾을 수 없습니다: {DATA_FOLDER}")
        return False
    
    print(f"📂 발견된 Excel 파일: {len(excel_files)}개\n")
    
    # 색상 맵
    color_map = {}
    
    # 주별 데이터 추출
    weeks = {}
    processed_count = 0
    
    SHEETS = list(SHEET_CONFIG.keys())
    
    for excel_file in excel_files:
        try:
            filename = os.path.basename(excel_file)
            week_code = get_week_code(filename)
            
            if not week_code:
                print(f"⚠️  {filename}: 주차 코드 추출 실패")
                continue
            
            print(f"📄 {filename}")
            print(f"   Week: {week_code}")
            
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            weeks[week_code] = {}
            
            for sheet_name in SHEETS:
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                sheet_data = extract_sheet_data(sheet, sheet_name, SHEET_CONFIG[sheet_name])
                weeks[week_code][sheet_name] = sheet_data
                
                # 색상 맵에 불량유형 추가
                for defect_type in sheet_data['defects'].keys():
                    if defect_type not in color_map:
                        idx = len(color_map)
                        color_map[defect_type] = DEFAULT_COLORS[idx % len(DEFAULT_COLORS)]
            
            workbook.close()
            processed_count += 1
            
        except Exception as e:
            print(f"❌ 오류: {e}")
            continue
    
    if processed_count == 0:
        print("\n❌ 처리된 파일이 없습니다.")
        return False
    
    # JSON 저장
    dashboard_data = {"weeks": weeks, "color_map": color_map}
    
    try:
        with open(OUTPUT_FILENAME, 'w', encoding='utf-8') as f:
            json.dump(dashboard_data, f, ensure_ascii=False, indent=2)
        
        file_size_kb = os.path.getsize(OUTPUT_FILENAME) / 1024
        unique_defects = len(color_map)
        
        print("\n" + "=" * 70)
        print("✅ 완료!")
        print("=" * 70)
        print(f"📊 처리 파일: {processed_count}개")
        print(f"📈 주간 데이터: {len(weeks)}주")
        print(f"🎨 불량유형: {unique_defects}가지")
        print(f"📁 출력 위치: {os.path.abspath(OUTPUT_FILENAME)}")
        print(f"📦 파일 크기: {file_size_kb:.1f} KB")
        print("\n🌐 Dashboard 접속:")
        print("   http://localhost:8000/dashboard.html")
        print("\n" + "=" * 70)
        
        return True
        
    except Exception as e:
        print(f"❌ 저장 오류: {e}")
        return False

if __name__ == "__main__":
    success = main()
    input("\n엔터를 눌러 종료하세요...")
    exit(0 if success else 1)
